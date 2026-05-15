import subprocess
import itertools
import openpyxl
import csv
import os
import numpy as np
import matplotlib.pyplot as plt

SRC = "triggerThreshold-arm2.cc"
OUT = "../bin/triggerThreshold-arm2"

# (hit, st, sw)
# (0,0,0): Miss Load
# (0,1,0): Miss Store
# (1,0,0): Hit Load
# (0,0,1): Miss Prefetch
# Miss Load
configs = [(0,0,0), (0,1,0), (1,0,0), (0,0,1)]

CPU_ID=6
micro_arch = f"Cortex-core{CPU_ID}"
wb = openpyxl.Workbook()

# 保存每个 sheet 对应的数据，后面画热图要用
heatmap_data = {}

for hit, st, sw in configs:
    print("=" * 60)
    # print(f"TEST_ON_HIT={hit}, TEST_ON_ST={st}, TEST_ON_SW={sw}")

    compile_cmd = [
        "g++",
        "-std=gnu11",
        "-O0",
        "-static",
        f"-DTEST_ON_HIT={hit}",
        f"-DTEST_ON_ST={st}",
        f"-DTEST_ON_SW={sw}",
        "-o",
        OUT,
        SRC
    ]

    res = subprocess.run(compile_cmd)

    if res.returncode != 0:
        print("Compile failed")
        continue

    run = subprocess.run(
        ["taskset", "-c", str(CPU_ID), "./" + OUT],
        capture_output=True,
        text=True
    )

    if run.returncode != 0:
        print("Execution failed")
        print(run.stderr)
        continue

    output = run.stdout.strip()
    if not output:
        print("Empty output")
        continue

    sheet_name = f"HIT{hit}-ST{st}-SW{sw}"
    ws = wb.create_sheet(sheet_name)

    lines = output.splitlines()

    # 第一行决定列数
    first = [x for x in lines[0].strip().split('\t') if x != ""]
    n = len(first)

    # Excel 表头
    ws.append(["Stride"] + [f"access{i}" for i in range(1, n + 1)])

    # 解析为矩阵，同时写入 Excel
    matrix = []
    matrix.append([float(x) for x in first])
    ws.append([1] + first)

    for idx, line in enumerate(lines[1:], start=2):
        cols = [x for x in line.strip().split('\t') if x != ""]
        if len(cols) == 0:
            continue
        ws.append([idx] + cols)
        matrix.append([float(x) for x in cols])

    heatmap_data[sheet_name] = np.array(matrix, dtype=float)

# 删除默认 sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

os.makedirs("res", exist_ok=True)

xlsx_path = f"res/threshold-{micro_arch}.xlsx"
wb.save(xlsx_path)
print(f"Excel saved to {xlsx_path}")

# =========================
# 画热图：1行4列
# =========================
plot_order = [
    "HIT0-ST0-SW0",  # Miss Load
    "HIT0-ST1-SW0",  # Miss Store
    "HIT1-ST0-SW0",  # Hit Load
    "HIT0-ST0-SW1",  # Miss Prefetch
]

plot_titles = {
    "HIT0-ST0-SW0": "Miss Load",
    "HIT0-ST1-SW0": "Miss Store",
    "HIT1-ST0-SW0": "Hit Load",
    "HIT0-ST0-SW1": "Miss Prefetch",
}

# 用统一颜色范围，便于横向比较
available_data = [heatmap_data[name] for name in plot_order if name in heatmap_data]
if not available_data:
    print("No data available for plotting.")
    exit(0)

# vmin = min(np.min(x) for x in available_data)
# vmax = max(np.max(x) for x in available_data)

fig, axes = plt.subplots(1, 4, figsize=(24, 5))

im = None
for i, name in enumerate(plot_order):
    ax = axes[i]

    if name not in heatmap_data:
        ax.set_title(f"{plot_titles[name]}\n(no data)")
        ax.axis("off")
        continue

    data = heatmap_data[name]

    im = ax.imshow(
        data,
        aspect='auto',
        vmin=20,
        vmax=400,
        cmap='RdYlBu_r'
    )

    ax.set_title(plot_titles[name], fontsize=12)
    ax.set_xlabel("train_step")
    ax.set_ylabel("stride")

    # 控制刻度密度，避免太挤
    if data.shape[1] > 1:
        ax.set_xticks(range(0, data.shape[1], max(1, data.shape[1] // 5)))
    if data.shape[0] > 1:
        ax.set_yticks(range(0, data.shape[0], max(1, data.shape[0] // 8)))

# 右侧统一 colorbar
cbar_ax = fig.add_axes([0.92, 0.15, 0.015, 0.7])
fig.colorbar(im, cax=cbar_ax)

plt.tight_layout(rect=[0, 0, 0.9, 1])

png_path = f"res/threshold-{micro_arch}-4in1.png"
plt.savefig(png_path, dpi=200)
print(f"Heatmap saved to {png_path}")

plt.show()

print("All done.")