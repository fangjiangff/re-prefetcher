import subprocess
import itertools
import openpyxl
import csv
import argparse

import os

# SRC = "triggerThreshold.cc"
ARCH_CONFIG = {
    "A76": {
        "core": 2,
        "src": "triggerThreshold-arm2.cc",
        "out": "bin/triggerThreshold-arm2",
    },
    "RaptorCove": {
        "core": 0,
        "src": "triggerThreshold-x86.cc",
        "out": "bin/triggerThreshold-x86",
    },
    "Gracemont": {
        "core": 16,
        "src": "triggerThreshold-x86.cc",
        "out": "bin/triggerThreshold-x86",
    },
}
DEFAULT_ARCH = "A76"
DEFAULT_STRIDE = 15
DEFAULT_OUTPUT_DIR = "res"
DEFAULT_HEATMAP_DIR = "res/heatmaps"

# configs = [(1,1), (1,0), (0,1), (0,0)]
# configs = [ (0,0), (1,0)]
# // (0,0,0)miss load, (0,1,0) miss store; (1,0,0) hit load,(0,0,1) miss prefetch.
# configs = [(0,0,0), (0,1,0), (1,0,0), (0,0,1)]
configs = [(0,0,0), (0,0,1)]
def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--stride", type=int, default=DEFAULT_STRIDE,
                        help=f"Stride in cache lines. Default: {DEFAULT_STRIDE}")
    parser.add_argument("--core", type=int, default=None,
                        help="Override CPU core used by taskset and CPU_ID. "
                             "Default is selected from --arch.")
    parser.add_argument("--arch", default=DEFAULT_ARCH, choices=ARCH_CONFIG.keys(),
                        help=f"Architecture name used in output filenames. Default: {DEFAULT_ARCH}")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR,
                        help=f"Directory for xlsx results. Default: {DEFAULT_OUTPUT_DIR}")
    parser.add_argument("--heatmap-dir", default=DEFAULT_HEATMAP_DIR,
                        help=f"Directory for generated heatmaps. Default: {DEFAULT_HEATMAP_DIR}")
    parser.add_argument("--plot-only", action="store_true",
                        help="Only plot from an existing result workbook.")
    args = parser.parse_args()
    if args.stride < 1:
        parser.error("--stride must be >= 1")
    if args.core is None:
        args.core = ARCH_CONFIG[args.arch]["core"]
    if args.core < 0:
        parser.error("--core must be >= 0")
    return args


args = parse_args()
micro_arch = f"{args.arch}-core{args.core}-stride={args.stride}"
wb = openpyxl.Workbook()
PLOT_ONLY = args.plot_only
INPUT_FILE = os.path.join(args.output_dir, f"threshold-{micro_arch}.xlsx")

TARGET_SHEETS = [
    "HIT0-ST0-SW0",  # Miss Load
    # "HIT0-ST1-SW0",  # Miss Store
    # "HIT1-ST0-SW0",  # Hit Load
    "HIT0-ST0-SW1",  # Miss Prefetch
]
title_map = {
    "HIT0-ST0-SW0": "Load instruction (miss)",
    # "HIT0-ST1-SW0": "Miss Store", 
    # "HIT1-ST0-SW0": "Hit Load",
    "HIT0-ST0-SW1": "Prefetch instruction (miss)"
}

def plot_heatmaps():
    try:
        import pandas as pd
        import seaborn as sns
        import matplotlib.pyplot as plt
    except ModuleNotFoundError as exc:
        print(f"Skipping heatmap: missing Python package '{exc.name}'.")
        print("Install plotting dependencies with:")
        print("  sudo apt install python3-pandas python3-openpyxl python3-seaborn python3-matplotlib")
        print("or, for the current Python environment:")
        print("  python3 -m pip install pandas openpyxl seaborn matplotlib")
        return

    # 检查输入文件是否存在
    if not os.path.exists(INPUT_FILE):
        print(f"Error: Input file '{INPUT_FILE}' not found.")
        return


    print(f"Reading data from {INPUT_FILE}...")
    try:
        # 加载 Excel 文件
        workbook = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)
    except Exception as e:
        print(f"Failed to open Excel file: {e}")
        return

    valid_sheets = []
    for sheet_name in TARGET_SHEETS:
        if sheet_name in workbook.sheetnames:
            valid_sheets.append(sheet_name)
        else:
            print(f"Warning: Sheet '{sheet_name}' not found in workbook. Skipping.")

    num_sheets = len(valid_sheets)
    if num_sheets == 0:
        print("No valid sheets found.")
        return

    # 创建等宽热图子图，并把 colorbar 放到单独的轴上，避免挤压最后一张热图。
    fig = plt.figure(figsize=(6 * num_sheets + 0.5, 8), constrained_layout=True)
    gs = fig.add_gridspec(
        1,
        num_sheets + 1,
        width_ratios=[1] * num_sheets + [0.04],
        wspace=0.08,
    )
    axes = [fig.add_subplot(gs[0, i]) for i in range(num_sheets)]
    cbar_ax = fig.add_subplot(gs[0, num_sheets])

    for i, sheet_name in enumerate(valid_sheets):
        print(f"Processing sheet: {sheet_name}")
        ax = axes[i]
        try:
            # 读取工作表数据
            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            if not rows:
                print(f"Warning: Sheet '{sheet_name}' is empty. Skipping.")
                continue
            df = pd.DataFrame(rows[1:], columns=rows[0])
            
            # 将第一列 'Stride' 设为索引
            if not df.empty:
                df.set_index(df.columns[0], inplace=True)
            df = df.apply(pd.to_numeric, errors="coerce")
            df = df.dropna(how="all").dropna(axis=1, how="all")

            # 绘制热力图
            sns.heatmap(
                df,
                cmap="RdYlBu_r",
                annot=False,
                ax=ax,
                vmin=0,
                vmax=400,
                cbar=(i == num_sheets - 1),
                cbar_ax=(cbar_ax if i == num_sheets - 1 else None),
                yticklabels=(i == 0),
                linewidths=0.35,
                linecolor='gray',
            )
            
            ax.set_title(f"{title_map.get(sheet_name, sheet_name)}", fontsize=14)
            ax.set_xlabel("Test Position (i * Stride)")
            ax.set_ylabel("Access Times (N)" if i == 0 else "")

        except Exception as e:
            print(f"Error processing {sheet_name}: {e}")
            ax.text(0.5, 0.5, "Error", ha='center', va='center')

    workbook.close()
    os.makedirs(args.heatmap_dir, exist_ok=True)
    output_path = os.path.join(args.heatmap_dir, f"{micro_arch}.png")
    plt.savefig(output_path, dpi=300)
    plt.close()
    print(f"Saved combined heatmap to {output_path}")
    print("All done.")


def run_tests():
    if PLOT_ONLY:
        return

    arch_config = ARCH_CONFIG[args.arch]
    src = arch_config["src"]
    out = arch_config["out"]
    os.makedirs(os.path.dirname(out), exist_ok=True)

    for hit,st, sw in configs:
        print("="*60)
        print(f"TEST_ON_HIT={hit}, TEST_ON_ST={st}, TEST_ON_SW={sw}, "
              f"ARCH={args.arch}, STRIDE={args.stride}, CORE={args.core}")
        # print(f"TEST_ON_HIT={hit}, TEST_ON_SW={sw}")

        compile_cmd = [
            "g++",
            "-std=gnu11",
            "-O0",
            "-static",
            f"-DTEST_ON_HIT={hit}",
            f"-DTEST_ON_ST={st}",
            f"-DTEST_ON_SW={sw}",
            f"-DSTRIDE={args.stride}",
            f"-DCPU_ID={args.core}",
            "-o",
            out,
            src
        ]

        res = subprocess.run(compile_cmd)

        if res.returncode != 0:
            print("Compile failed")
            continue

        run = subprocess.run(
            ["taskset", "-c", str(args.core), "./" + out],
            capture_output=True,
            text=True
        )

        if run.returncode != 0:
            print("Execution failed")
            continue

        output = run.stdout
        lines = output.splitlines()
        if not lines:
            print("Execution produced no output")
            continue

        sheet_name = f"HIT{hit}-ST{st}-SW{sw}"
        ws = wb.create_sheet(sheet_name)

        # 读第一行，确定列数
        # first = output.readline().strip().split('\t')
        first = lines[0].strip().split('\t')
        n = len(first)
        # 写表头
        ws.append(["Stride"] + [f"pos{i}" for i in range(0, n)])

        # 写第一行数据
        ws.append([1] + first)

        # 写后续行
        # for idx, line in enumerate(output, start=2):
        #     ws.append([idx] + line.strip().split('\t'))
        # 修复后代码
        for idx, line in enumerate(lines[1:], start=2):
            ws.append([idx] + line.strip().split('\t'))

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    os.makedirs(args.output_dir, exist_ok=True)
    wb.save(INPUT_FILE)


if __name__ == "__main__":
    run_tests()
    plot_heatmaps()
# print("All done.")
