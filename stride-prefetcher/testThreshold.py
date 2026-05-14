import subprocess
import itertools
import openpyxl
import csv
import argparse

import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt

# SRC = "triggerThreshold.cc"
# SRC = "triggerThreshold-x86.cc"
# OUT = "bin/triggerThreshold-x86"
SRC = "triggerThreshold-arm.cc"
OUT = "../bin/triggerThreshold-arm"

# configs = [(1,1), (1,0), (0,1), (0,0)]
# configs = [ (0,0), (1,0)]
# // (0,0,0)miss load, (0,1,0) miss store; (1,0,0) hit load,(0,0,1) miss prefetch.
# configs = [(0,0,0), (0,1,0), (1,0,0), (0,0,1)]
configs = [(0,0,0), (0,0,1)]
DEFAULT_ARCH = "CortexA78"
DEFAULT_CORE = 5
DEFAULT_PROBE_POSITION = 1

def ordinal(n):
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--probe-position", type=int, default=DEFAULT_PROBE_POSITION,
                        help=f"Probe the Nth expected prefetched position. Default: {DEFAULT_PROBE_POSITION}")
    parser.add_argument("--core", type=int, default=DEFAULT_CORE,
                        help=f"CPU core used by taskset and CPU_ID. Default: {DEFAULT_CORE}")
    parser.add_argument("--arch", default=DEFAULT_ARCH,
                        help=f"Architecture name used in output filenames. Default: {DEFAULT_ARCH}")
    parser.add_argument("--plot-only", action="store_true",
                        help="Only plot from an existing result workbook.")
    args = parser.parse_args()
    if args.probe_position < 1:
        parser.error("--probe-position must be >= 1")
    if args.core < 0:
        parser.error("--core must be >= 0")
    return args

args = parse_args()
# micro_arch = "CascadeLake"
micro_arch = f"{args.arch}-core-{args.core}-probe-{ordinal(args.probe_position)}"
wb = openpyxl.Workbook()
PLOT_ONLY = args.plot_only
# timestamp = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
# print(f"Start testing at {timestamp}")
def run_tests():
    if PLOT_ONLY:
        return

    for hit,st, sw in configs:
        print("="*60)
        print(f"TEST_ON_HIT={hit}, TEST_ON_ST={st}, TEST_ON_SW={sw}, "
              f"PROBE_POSITION={args.probe_position}, CORE={args.core}")
        # print(f"TEST_ON_HIT={hit}, TEST_ON_SW={sw}")

        compile_cmd = [
            "g++",
            "-std=gnu11",
            "-O0",
            "-static",
            f"-DTEST_ON_HIT={hit}",
            f"-DTEST_ON_ST={st}",
            f"-DTEST_ON_SW={sw}",
            f"-DPROBE_POSITION={args.probe_position}",
            f"-DCPU_ID={args.core}",
            "-o",
            OUT,
            SRC
        ]

        res = subprocess.run(compile_cmd)

        if res.returncode != 0:
            print("Compile failed")
            continue

        run = subprocess.run(
            ["taskset", "-c", str(args.core), "./" + OUT],
            capture_output=True,
            text=True
        )

        if run.returncode != 0:
            print("Execution failed")
            print(run.stderr)
            continue

        output = run.stdout
        lines = output.splitlines()
        if not lines:
            print("Execution produced no output")
            continue

        sheet_name = f"HIT{hit}-ST{st}-SW{sw}"
        ws = wb.create_sheet(sheet_name)


        # 读第一行，确定列数
        # first = output.readline().strip().split('\t')
        first = lines[0].strip().split('\t')
        n = len(first)
        # 写表头
        ws.append(["Stride"] + [f"pos{i}" for i in range(1, n+1)])

        # 写第一行数据
        ws.append([1] + first)

        # 写后续行
        # for idx, line in enumerate(output, start=2):
        #     ws.append([idx] + line.strip().split('\t'))
        # 修复后代码
        for idx, line in enumerate(lines[1:], start=2):
            ws.append([idx] + line.strip().split('	'))


    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(f"res/threshold-{micro_arch}.xlsx")


INPUT_FILE = f"res/threshold-{micro_arch}.xlsx"

TARGET_SHEETS = [
    "HIT0-ST0-SW0",  # Miss Load
    # "HIT0-ST1-SW0",  # Miss Store
    # "HIT1-ST0-SW0",  # Hit Load
    "HIT0-ST0-SW1",  # Miss Prefetch
]

def plot_heatmaps():
    # 检查输入文件是否存在
    if not os.path.exists(INPUT_FILE):
        print(f"Error: Input file '{INPUT_FILE}' not found.")
        return


    print(f"Reading data from {INPUT_FILE}...")
    try:
        # 加载 Excel 文件
        xls = pd.ExcelFile(INPUT_FILE)
    except Exception as e:
        print(f"Failed to open Excel file: {e}")
        return

    valid_sheets = []
    for sheet_name in TARGET_SHEETS:
        if sheet_name in xls.sheet_names:
            valid_sheets.append(sheet_name)
        else:
            print(f"Warning: Sheet '{sheet_name}' not found in workbook. Skipping.")

    num_sheets = len(valid_sheets)
    if num_sheets == 0:
        print("No valid sheets found.")
        return

    # 创建子图：1行 num_sheets 列
    fig, axes = plt.subplots(1, num_sheets, figsize=(6 * num_sheets, 8))
    if num_sheets == 1:
        axes = [axes]

    for i, sheet_name in enumerate(valid_sheets):
        print(f"Processing sheet: {sheet_name}")
        ax = axes[i]
        try:
            # 读取工作表数据
            df = pd.read_excel(xls, sheet_name=sheet_name)
            
            # 将第一列 'Stride' 设为索引
            if not df.empty:
                df.set_index(df.columns[0], inplace=True)

            # 绘制热力图
            sns.heatmap(df, cmap="viridis", annot=False, ax=ax, vmin=20, vmax=600, cbar=(i == num_sheets - 1), yticklabels=(i == 0))
            
            ax.set_title(f"{sheet_name}")
            ax.set_xlabel("Access Sequence")
            ax.set_ylabel("Stride / Index" if i == 0 else "")

        except Exception as e:
            print(f"Error processing {sheet_name}: {e}")
            ax.text(0.5, 0.5, "Error", ha='center', va='center')

    plt.tight_layout()
    output_path = os.path.join("res/heatmaps", f"{micro_arch}.png")
    plt.savefig(output_path, dpi=300)
    plt.close()
    print(f"Saved combined heatmap to {output_path}")
    print("All done.")


if __name__ == "__main__":
    run_tests()
    plot_heatmaps()
# print("All done.")
