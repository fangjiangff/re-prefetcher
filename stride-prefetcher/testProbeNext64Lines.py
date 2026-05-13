import argparse
import os
import subprocess
import sys

import openpyxl


SRC = "triggerThreshold-arm2.cc"
OUT = "bin/triggerThreshold-arm2"
DEFAULT_ARCH = "CortexA76"
DEFAULT_MIN_STRIDE = 1
DEFAULT_MAX_STRIDE = 29
DEFAULT_CORE = 0
DEFAULT_OUTPUT_DIR = "res/probeNext64Lines"
DEFAULT_ITEMS = 4096
DEFAULT_PROBE_POSITIONS = 64

# (0,0,0) miss load; (0,0,1) miss prefetch.
CONFIGS = [(0, 0, 0), (0, 0, 1)]
TITLE_MAP = {
    "HIT0-ST0-SW0": "Load instruction (miss)",
    "HIT0-ST0-SW1": "Prefetch instruction (miss)",
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Sweep STRIDE from 1 to 64 and plot probe-next-line heatmaps."
    )
    parser.add_argument("--arch", default=DEFAULT_ARCH)
    parser.add_argument("--min-stride", type=int, default=DEFAULT_MIN_STRIDE)
    parser.add_argument("--max-stride", type=int, default=DEFAULT_MAX_STRIDE)
    parser.add_argument("--core", type=int, default=DEFAULT_CORE)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--items", type=int, default=DEFAULT_ITEMS)
    parser.add_argument("--probe-positions", type=int, default=DEFAULT_PROBE_POSITIONS)
    parser.add_argument("--plot-vmin", type=float, default=20)
    parser.add_argument("--plot-vmax", type=float, default=400)
    parser.add_argument("--plot-only", action="store_true")
    parser.add_argument("--no-plot", action="store_true")
    return parser.parse_args()


def compile_test(args, hit, st, sw, stride):
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    compile_cmd = [
        "g++",
        "-std=gnu11",
        "-O0",
        "-static",
        f"-DTEST_ON_HIT={hit}",
        f"-DTEST_ON_ST={st}",
        f"-DTEST_ON_SW={sw}",
        f"-DSTRIDE={stride}",
        f"-DItems={args.items}",
        f"-DPROBE_POSITIONS={args.probe_positions}",
        "-o",
        OUT,
        SRC,
    ]
    return subprocess.run(compile_cmd)


def run_test(core):
    run_cmd = ["taskset", "-c", str(core), "./" + OUT]
    return subprocess.run(run_cmd, capture_output=True, text=True)


def parse_output(output):
    rows = []
    for line in output.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        rows.append([float(value) for value in stripped.split()])

    if not rows:
        raise ValueError("execution produced no plottable output")

    width = len(rows[0])
    if any(len(row) != width for row in rows):
        raise ValueError("execution output has inconsistent row widths")

    return rows


def append_matrix(ws, rows):
    width = len(rows[0])
    ws.append(["TrainStep"] + [f"pos{i}" for i in range(width)])
    for idx, row in enumerate(rows, start=1):
        ws.append([idx] + row)


def write_workbook(path, stride_results):
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    for stride, rows in stride_results:
        ws = wb.create_sheet(f"stride_{stride}")
        append_matrix(ws, rows)
    wb.save(path)


def plot_stride_heatmap(path, rows, title, vmin, vmax):
    import matplotlib.pyplot as plt
    import numpy as np
    import seaborn as sns

    data = np.array(rows, dtype=float)
    fig, ax = plt.subplots(figsize=(12, 8))
    sns.heatmap(
        data,
        cmap="viridis",
        annot=False,
        ax=ax,
        vmin=vmin,
        vmax=vmax,
        linewidths=0.25,
        linecolor="gray",
    )
    ax.set_title(title, fontsize=14)
    ax.set_xlabel("Probe Position (pos * stride)")
    ax.set_ylabel("Train Step")
    ax.set_xticks([i + 0.5 for i in range(data.shape[1])])
    ax.set_yticks([i + 0.5 for i in range(data.shape[0])])
    ax.set_xticklabels([str(i) for i in range(data.shape[1])], rotation=90)
    ax.set_yticklabels([str(i) for i in range(1, data.shape[0] + 1)], rotation=0)
    fig.tight_layout()
    fig.savefig(path, dpi=300)
    plt.close(fig)


def plot_probe_next_summary(path, stride_results, title, vmin, vmax):
    import matplotlib.pyplot as plt
    import numpy as np
    import seaborn as sns

    summary_rows = []
    stride_labels = []
    max_train_steps = 0
    for stride, rows in stride_results:
        next_line_values = []
        for train_step, row in enumerate(rows, start=1):
            next_pos = train_step
            next_line_values.append(row[next_pos] if next_pos < len(row) else float("nan"))
        summary_rows.append(next_line_values)
        stride_labels.append(str(stride))
        max_train_steps = max(max_train_steps, len(next_line_values))

    data = np.array(summary_rows, dtype=float)
    fig, ax = plt.subplots(figsize=(12, 12))
    sns.heatmap(
        data,
        cmap="viridis",
        annot=False,
        ax=ax,
        vmin=vmin,
        vmax=vmax,
        linewidths=0.25,
        linecolor="gray",
    )
    ax.set_title(title, fontsize=14)
    ax.set_xlabel("Train Step; value probes pos=train_step")
    ax.set_ylabel("Stride (cache lines)")
    ax.set_xticks([i + 0.5 for i in range(max_train_steps)])
    ax.set_yticks([i + 0.5 for i in range(len(stride_labels))])
    ax.set_xticklabels([str(i) for i in range(1, max_train_steps + 1)], rotation=90)
    ax.set_yticklabels(stride_labels, rotation=0)
    fig.tight_layout()
    fig.savefig(path, dpi=300)
    plt.close(fig)


def plot_results(args, config_name, stride_results):
    os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-cache")

    heatmap_dir = os.path.join(args.output_dir, "heatmaps", config_name)
    os.makedirs(heatmap_dir, exist_ok=True)

    for stride, rows in stride_results:
        path = os.path.join(heatmap_dir, f"{args.arch}-{config_name}-stride={stride}.png")
        title = f"{TITLE_MAP.get(config_name, config_name)}; stride={stride}"
        plot_stride_heatmap(path, rows, title, args.plot_vmin, args.plot_vmax)
        print(f"Saved heatmap to {path}")

    summary_path = os.path.join(args.output_dir, f"{args.arch}-{config_name}-probe-next-summary.png")
    summary_title = f"{TITLE_MAP.get(config_name, config_name)}; probe next line summary"
    plot_probe_next_summary(
        summary_path,
        stride_results,
        summary_title,
        args.plot_vmin,
        args.plot_vmax,
    )
    print(f"Saved summary heatmap to {summary_path}")


def run_sweep(args):
    if args.min_stride < 1 or args.max_stride < args.min_stride:
        raise ValueError("invalid stride range")

    os.makedirs(args.output_dir, exist_ok=True)
    all_results = {}

    for hit, st, sw in CONFIGS:
        config_name = f"HIT{hit}-ST{st}-SW{sw}"
        stride_results = []

        for stride in range(args.min_stride, args.max_stride + 1):
            print("=" * 60)
            print(f"{config_name}, STRIDE={stride}")

            res = compile_test(args, hit, st, sw, stride)
            if res.returncode != 0:
                raise RuntimeError(f"compile failed for {config_name}, stride={stride}")

            run = run_test(args.core)
            if run.stderr:
                print(run.stderr, end="", file=sys.stderr)
            if run.returncode != 0:
                raise RuntimeError(f"execution failed for {config_name}, stride={stride}")

            rows = parse_output(run.stdout)
            stride_results.append((stride, rows))

        result_path = os.path.join(args.output_dir, f"{args.arch}-{config_name}.xlsx")
        write_workbook(result_path, stride_results)
        print(f"Saved result to {result_path}")
        all_results[config_name] = stride_results

    return all_results


def load_workbook_results(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    stride_results = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("stride_"):
            continue
        stride = int(sheet_name.split("_", 1)[1])
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append([float(value) for value in row[1:]])
        stride_results.append((stride, rows))
    return sorted(stride_results, key=lambda item: item[0])


def main():
    args = parse_args()

    try:
        if args.plot_only:
            all_results = {}
            for hit, st, sw in CONFIGS:
                config_name = f"HIT{hit}-ST{st}-SW{sw}"
                path = os.path.join(args.output_dir, f"{args.arch}-{config_name}.xlsx")
                all_results[config_name] = load_workbook_results(path)
        else:
            all_results = run_sweep(args)

        if not args.no_plot:
            for config_name, stride_results in all_results.items():
                plot_results(args, config_name, stride_results)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print("All done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
