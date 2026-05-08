import subprocess
import openpyxl
import os
import numpy as np
import matplotlib.pyplot as plt

SRC = "triggerThreshold-arm.cc"
OUT = "../bin/triggerThreshold-arm"

TEST_ON_SW = 1

PRFM_MODES = [
    "PLDL1KEEP",
    "PSTL1KEEP",
    "PLDL2KEEP",
    "PSTL2KEEP",
    "PLDL3KEEP",
    "PSTL3KEEP",
    "PLDL1STRM",
    "PSTL1STRM",
    "PLDL2STRM",
    "PSTL2STRM",
    "PLDL3STRM",
    "PSTL3STRM",
]

CPU_ID=0
micro_arch = f"Cortex-core{CPU_ID}-disable-sti-sts-region-pf"

wb = openpyxl.Workbook()

# 用于绘图的数据
heatmap_data = {}

for mode in PRFM_MODES:
    print("=" * 60)
    print(f"TEST_ON_SW=1, PRFM_MODE={mode}")

    compile_cmd = [
        "g++",
        "-std=gnu++17",
        "-O0",
        f"-DTEST_ON_SW={TEST_ON_SW}",
        f"-DPRFM_MODE={mode}",
        "-o",
        OUT,
        SRC,
    ]

    res = subprocess.run(compile_cmd)
    if res.returncode != 0:
        print("Compile failed")
        continue

    run = subprocess.run(
        ["taskset", "-c", str(CPU_ID), "./" + OUT],
        capture_output=True,
        text=True
    )

    if run.returncode != 0:
        print("Execution failed")
        print(run.stderr)
        continue

    output = run.stdout.strip()
    if not output:
        print("Empty output")
        continue

    lines = output.splitlines()

    # 解析数据
    matrix = []
    for line in lines:
        row = [float(x) for x in line.strip().split('\t') if x != ""]
        matrix.append(row)

    matrix = np.array(matrix)
    heatmap_data[mode] = matrix

    # 写 Excel
    ws = wb.create_sheet(mode[:31])
    n = matrix.shape[1]

    ws.append(["Stride"] + [f"access{i}" for i in range(1, n + 1)])

    for i, row in enumerate(matrix, start=1):
        ws.append([i] + list(row))

# 删除默认 sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

os.makedirs("res", exist_ok=True)
xlsx_path = f"res/sw-{micro_arch}.xlsx"
wb.save(xlsx_path)

print(f"Excel saved to {xlsx_path}")

# =============================
# 🔥 画热图（2 x 6 + 右侧 colorbar）
# =============================

fig, axes = plt.subplots(2, 6, figsize=(24, 8))

# 固定范围（关键）
vmin = 50
vmax = 300

im = None

for idx, mode in enumerate(PRFM_MODES):
    row = idx // 6
    col = idx % 6

    ax = axes[row][col]
    data = heatmap_data[mode]

    im = ax.imshow(
        data,
        aspect='auto',
        vmin=vmin,
        vmax=vmax,
        cmap='RdYlBu_r'   # 推荐：对 latency 很直观
    )

    ax.set_title(mode, fontsize=10)

    ax.set_xlabel("train_step")
    ax.set_ylabel("stride")

    # 控制刻度密度
    ax.set_xticks(range(0, data.shape[1], 5))
    ax.set_yticks(range(0, data.shape[0], 5))

# 👉 在最右侧单独放 colorbar
# 用 add_axes 精确控制位置
cbar_ax = fig.add_axes([0.92, 0.15, 0.015, 0.7])  
# 参数含义：[left, bottom, width, height]

fig.colorbar(im, cax=cbar_ax)

plt.tight_layout(rect=[0, 0, 0.9, 1])  
# 留出右侧空间给 colorbar

png_path = f"res/sw-{micro_arch}.png"
plt.savefig(png_path, dpi=200)

print(f"Heatmap saved to {png_path}")

plt.show()